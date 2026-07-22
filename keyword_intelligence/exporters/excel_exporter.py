"""Excel exporter for comprehensive reports."""

from pathlib import Path

import pandas as pd
from loguru import logger

from keyword_intelligence.exporters.base import BaseExporter
from keyword_intelligence.models import PipelineResult
from keyword_intelligence.pipeline.context import PipelineContext


class ExcelExporter(BaseExporter):
    """Generates the formatted filtered_keywords.xlsx artifact."""

    def export(
        self, result: PipelineResult, context: PipelineContext, output_dir: Path
    ) -> list[str]:
        if not context.has_data:
            return []

        xlsx_path = output_dir / "filtered_keywords.xlsx"
        df = context.data

        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                # 1. Final Results (Clean Business Columns)
                df_business = df.rename(
                    columns={
                        "classification_stage": "classification",
                        "decision_confidence": "confidence",
                        "company_name": "company",
                        "company_website": "website",
                    }
                ).copy()

                # Map boolean
                if "relevant" in df_business.columns:
                    df_business["relevant"] = df_business["relevant"].map(
                        {
                            True: "Relevant",
                            False: "Irrelevant",
                            "True": "Relevant",
                            "False": "Irrelevant",
                        }
                    )

                target_cols = [
                    "keyword",
                    "search_volume",
                    "relevant",
                    "relevance_score",
                    "brand",
                    "category",
                    "classification",
                    "confidence",
                    "reason",
                    "recommended_action",
                    "company",
                    "website",
                ]
                available_cols = [c for c in target_cols if c in df_business.columns]
                df_business[available_cols].to_excel(
                    writer, sheet_name="Final Results", index=False
                )

                final_sheet = writer.sheets["Final Results"]
                self._format_sheet(final_sheet)

                # Conditional Formatting for Final Results
                from openpyxl.formatting.rule import CellIsRule
                from openpyxl.styles import PatternFill

                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                red_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
                orange_fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )

                if final_sheet.max_row > 1:
                    # formatting rules applied to entire data range for simplicity, or specific columns.
                    range_str = f"A2:Z{final_sheet.max_row}"
                    final_sheet.conditional_formatting.add(
                        range_str,
                        CellIsRule(
                            operator="equal", formula=['"Relevant"'], fill=green_fill
                        ),
                    )
                    final_sheet.conditional_formatting.add(
                        range_str,
                        CellIsRule(
                            operator="equal", formula=['"Irrelevant"'], fill=red_fill
                        ),
                    )
                    final_sheet.conditional_formatting.add(
                        range_str,
                        CellIsRule(
                            operator="equal",
                            formula=['"AI Classification"'],
                            fill=orange_fill,
                        ),
                    )

                # 2. Pipeline Summary
                summary_data = {
                    "Metric": [
                        "Total Keywords",
                        "Relevant",
                        "Irrelevant",
                        "Deterministic",
                        "AI",
                        "Duplicates Removed",
                        "Runtime (s)",
                        "AI Calls Saved",
                    ],
                    "Value": [
                        len(df),
                        int(df["relevant"].sum()) if "relevant" in df.columns else 0,
                        len(df) - int(df["relevant"].sum())
                        if "relevant" in df.columns
                        else len(df),
                        int((df["processing_method"] == "Deterministic").sum())
                        if "processing_method" in df.columns
                        else 0,
                        int((df["processing_method"] == "AI").sum())
                        if "processing_method" in df.columns
                        else 0,
                        result.metrics.total_rows_removed,
                        round(result.total_execution_time_ms / 1000.0, 2),
                        len(df)
                        - (
                            int((df["processing_method"] == "AI").sum())
                            if "processing_method" in df.columns
                            else 0
                        ),
                    ],
                }
                pd.DataFrame(summary_data).to_excel(
                    writer, sheet_name="Pipeline Summary", index=False
                )

                summary_sheet = writer.sheets["Pipeline Summary"]
                self._format_sheet(summary_sheet)

                # Add Charts to Pipeline Summary
                from openpyxl.chart import BarChart, PieChart, Reference

                if summary_sheet.max_row >= 6:
                    pie = PieChart()
                    pie.title = "Relevant vs Irrelevant"
                    data = Reference(summary_sheet, min_col=2, min_row=2, max_row=3)
                    labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=3)
                    pie.add_data(data, titles_from_data=False)
                    pie.set_categories(labels)
                    summary_sheet.add_chart(pie, "D2")

                    bar = BarChart()
                    bar.title = "Processing Method (Deterministic vs AI)"
                    bar.y_axis.title = "Keywords"
                    data = Reference(summary_sheet, min_col=2, min_row=4, max_row=5)
                    labels = Reference(summary_sheet, min_col=1, min_row=4, max_row=5)
                    bar.add_data(data, titles_from_data=False)
                    bar.set_categories(labels)
                    summary_sheet.add_chart(bar, "D18")

                # 3. Stage Metrics
                stage_data = [
                    {
                        "Stage": m.stage_name,
                        "Duration (ms)": round(m.processing_time_ms, 2),
                        "Input Rows": m.rows_loaded,
                        "Output Rows": m.rows_output,
                    }
                    for m in result.stage_metrics
                ]
                pd.DataFrame(stage_data).to_excel(
                    writer, sheet_name="Stage Metrics", index=False
                )
                self._format_sheet(writer.sheets["Stage Metrics"])

                # 4. AI Decisions
                if "requires_ai" in df.columns:
                    ai_df = df[df["requires_ai"] == True].copy()
                    if not ai_df.empty:
                        ai_cols = ["keyword", "relevant", "ai_confidence", "ai_reason"]
                        ai_cols = [c for c in ai_cols if c in ai_df.columns]
                        if not ai_cols:
                            ai_cols = ["keyword"]
                        ai_df[ai_cols].to_excel(
                            writer, sheet_name="AI Decisions", index=False
                        )
                        self._format_sheet(writer.sheets["AI Decisions"])
                    else:
                        pd.DataFrame(
                            columns=[
                                "keyword",
                                "relevant",
                                "ai_confidence",
                                "ai_reason",
                            ]
                        ).to_excel(writer, sheet_name="AI Decisions", index=False)
                        self._format_sheet(writer.sheets["AI Decisions"])

                # 5. Debug Data
                df.to_excel(writer, sheet_name="Debug Data", index=False)
                self._format_sheet(writer.sheets["Debug Data"])

            return [str(xlsx_path.absolute())]
        except Exception as e:
            logger.error(f"Failed to generate Excel output: {e}")
            return []

    def _format_sheet(self, worksheet):
        """Apply bold headers, auto-width columns, freeze first row, and auto filters."""
        from openpyxl.styles import Font

        # Freeze first row
        worksheet.freeze_panes = "A2"

        # Format Headers and Auto Filters
        if worksheet.max_row > 0:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

        # Auto-width columns
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = min(adjusted_width, 100)
